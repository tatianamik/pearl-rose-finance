import openpyxl

wb = openpyxl.load_workbook('/Users/tanyahughes/Downloads/agd-budgets/Alpha Omicron 2026-2027 Budget.xlsx', data_only=True)

ws = wb['Summary']
print("\n--- Sheet: Summary Expenses (50-100) ---")
for row in range(50, 80):
    row_vals = []
    for col in range(5, 8):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            row_vals.append(f"{openpyxl.utils.get_column_letter(col)}{row}: {val}")
    if row_vals:
        print(" | ".join(row_vals))

